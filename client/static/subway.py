# -*- coding: utf-8 -*-
# @Author: wakouboy
# @Date:   2018-08-27 15:48:05
# @Last Modified by:   wakouboy
# @Last Modified time: 2018-08-27 19:36:09
import json
from openpyxl import load_workbook
wb = load_workbook(filename = 'subway.xlsx')
sheetname = ['1', '2', '4', '5', '6', '7', '8', '9', '10', '13', '14', '142', '15', '16', '八通', '昌平', 
'亦庄', '房山', '机场', '西郊']

nodes = []
links = []
rowCnt = 100
allStations = []
for name in sheetname:
    sheet = wb[name]
    stations = []
    for row in range(1, rowCnt):
        station = sheet.cell(row = row, column = 1).value
        if station != '' and station != None:
            stations.append(station)
    # print(stations)
    for station in stations:
        if not station in allStations:
            allStations.append(station)
            nodes.append({'id': station})
    for i in range(len(stations)-1):
        links.append({'source': stations[i], 'target': stations[i+1]})

# print(nodes)
# print(links)
graph = {'nodes': nodes, 'links': links}

with open('subway.json', 'w') as f:
    json.dump(graph, f)


